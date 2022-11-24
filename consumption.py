#################################
#################################
#### group_18_ransomware.py#######
#### Group 18###Prof. Tunc###UNT##
#### CSCE 5550####FA22############
#################################
#################################
# Disclaimer: Ransomware is ILLEGAL. Do not run in production environments.
# This ransomware program was built for educational purposes as part of a research project. If you use
# this then it is at your own risk.
# This code is for monitoring the virus in a secure environment
# Author: Junaidsheik
# Version 1.1 11/18/2022
#######################################################################################################################
#######################################################################################################################
# Instructions:
#
# Step 1: Run the file consumption.py
#
# Step 2: Provide the specific Process ID which you want to monitor.
#
# Step 3: All the logs will be saved in to the .xlsx format.
#
# Step 4: Provide the correct location of the file.
#######################################################################################################################

import psutil
import datetime
import time
import schedule
import openpyxl
import os

pid = int(input("Enter Process ID: "))


def Alert():
    cpuusage = psutil.cpu_percent(interval=1)
    if cpuusage > 1:  # CPU consumption percentage
        print("Alert! CPU consumption is high", cpuusage)
    memusage = psutil.virtual_memory( ).percent
    if memusage > 50:  # Mem consumption percentage
        print("Memory consumption is high", memusage)


def monitor():
    time = datetime.datetime.now( ).strftime("%Y%m%d - %H:%M:%S")
    p = psutil.Process(pid)
    cpu = p.cpu_percent(interval=1) / psutil.cpu_count( )

    memory_mb = p.memory_full_info( ).rss / (1024 * 1024)
    memory = p.memory_percent( )

    path = r"C:\Users\Jshei\Desktop\Monitor_Result.xlsx"  # filepath
    assert os.path.isfile(path)
    with open(path, "r") as f:
        pass
    file = openpyxl.load_workbook(path)
    sheet = file.active

    sheet.cell(column=1, row=sheet.max_row + 1, value=time)
    sheet.cell(column=2, row=sheet.max_row, value=pid)
    sheet.cell(column=3, row=sheet.max_row, value=cpu)
    sheet.cell(column=4, row=sheet.max_row, value=memory_mb)
    sheet.cell(column=5, row=sheet.max_row, value=memory)
    file.save(path)


schedule.every(1).second.do(Alert)
schedule.every(1).second.do(monitor)

while True:
    schedule.run_pending( )
    time.sleep(1)
